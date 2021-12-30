import openpyxl
from openpyxl import Workbook, load_workbook

# wb = load_workbook('test_dataset.xlsx')
# ws = wb.active

# # Access cell in excel file
# print(ws['A2'].value)

# # Change cell value
# ws['A1'].value = 'Hello cho meo'
# wb.save('test_dataset.xlsx')

# # Access all sheets
# sheets = wb.sheetnames
# print(type(sheets))

# Creat new workbook
wb = Workbook()
ws = wb.active
ws.title = 'My data'
title_list = ['Cho', 'Meo', 'Ga tay', 'Da dieu']
ws.append(title_list)

wb.save('My_test.xlsx')